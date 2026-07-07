import pytest
import openpyxl
from pathlib import Path
from sheetci.scanner import (
    WorkbookScanner,
    RULE_BROKEN_REF,
    RULE_EXTERNAL_LINK,
    RULE_HIDDEN_SHEET,
    RULE_HARDCODED_NUMBER,
    RULE_SELF_REFERENCE,
    RULE_INCONSISTENT_FORMULA,
    RULE_CACHED_ERROR,
    normalize_formula,
    extract_hardcoded_numbers
)

def test_normalize_formula():
    # Relative cell reference in row 2
    assert normalize_formula("=A2*B2", 2) == "=A[0]*B[0]"
    # Absolute cell references should not normalize row offset
    assert normalize_formula("=$A$2*B$2", 2) == "=$A$2*B$2"
    # Mixed relative/absolute reference in row 5
    assert normalize_formula("=A5+C$1", 5) == "=A[0]+C$1"
    # Relative rows with offset in row 5
    assert normalize_formula("=A4+A6", 5) == "=A[-1]+A[1]"

def test_extract_hardcoded_numbers():
    assert extract_hardcoded_numbers("=A1*0.08") == [0.08]
    # Ignores harmless numbers 0, 1, 2
    assert extract_hardcoded_numbers("=SUM(A1:A10, 1, 0, -2)") == []
    # Test formula with strings and functions containing numbers
    assert extract_hardcoded_numbers('=IF(DEC2HEX(A1)="1A", 100, 2)') == [100.0]

def test_broken_ref_detector(tmp_path):
    wb_path = tmp_path / "broken_ref.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "=B1+#REF!"
    wb.save(wb_path)
    
    scanner = WorkbookScanner(str(wb_path))
    res = scanner.scan()
    
    findings = [f for f in res["findings"] if f["rule_id"] == RULE_BROKEN_REF]
    assert len(findings) == 1
    assert findings[0]["cell_address"] == "A1"
    assert findings[0]["severity"] == "critical"

def test_external_link_detector(tmp_path):
    wb_path = tmp_path / "external_link.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "=[Budget.xlsx]Sheet1!B2"
    wb.save(wb_path)
    
    scanner = WorkbookScanner(str(wb_path))
    res = scanner.scan()
    
    findings = [f for f in res["findings"] if f["rule_id"] == RULE_EXTERNAL_LINK]
    assert len(findings) == 1
    assert findings[0]["cell_address"] == "A1"
    assert findings[0]["severity"] == "warning"

def test_hidden_sheet_detector(tmp_path):
    wb_path = tmp_path / "hidden_sheet.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Visible"
    hidden_ws = wb.create_sheet("Hidden")
    hidden_ws.sheet_state = "hidden"
    wb.save(wb_path)
    
    scanner = WorkbookScanner(str(wb_path))
    res = scanner.scan()
    
    findings = [f for f in res["findings"] if f["rule_id"] == RULE_HIDDEN_SHEET]
    assert len(findings) == 1
    assert findings[0]["sheet_name"] == "Hidden"
    assert findings[0]["severity"] == "warning"

def test_hardcoded_number_detector(tmp_path):
    wb_path = tmp_path / "hardcoded_number.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "=B1*123.45"
    wb.save(wb_path)
    
    scanner = WorkbookScanner(str(wb_path))
    res = scanner.scan()
    
    findings = [f for f in res["findings"] if f["rule_id"] == RULE_HARDCODED_NUMBER]
    assert len(findings) == 1
    assert findings[0]["cell_address"] == "A1"
    assert findings[0]["severity"] == "warning"

def test_self_reference_detector(tmp_path):
    wb_path = tmp_path / "self_ref.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "=A1+10"
    wb.save(wb_path)
    
    scanner = WorkbookScanner(str(wb_path))
    res = scanner.scan()
    
    findings = [f for f in res["findings"] if f["rule_id"] == RULE_SELF_REFERENCE]
    assert len(findings) == 1
    assert findings[0]["cell_address"] == "A1"
    assert findings[0]["severity"] == "critical"

def test_inconsistent_formula_detector(tmp_path):
    wb_path = tmp_path / "inconsistent.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    # Set up 5 identical relative formulas in column C and 1 inconsistent formula
    for r in range(2, 7):
        ws[f"A{r}"] = r
        ws[f"B{r}"] = 10
        ws[f"C{r}"] = f"=A{r}*B{r}"
    ws["C7"] = "=A7*B7+100"  # Inconsistent!
    wb.save(wb_path)
    
    scanner = WorkbookScanner(str(wb_path))
    res = scanner.scan()
    
    findings = [f for f in res["findings"] if f["rule_id"] == RULE_INCONSISTENT_FORMULA]
    assert len(findings) == 1
    assert findings[0]["cell_address"] == "C7"
    assert findings[0]["severity"] == "warning"

def test_cached_error_detector(tmp_path):
    wb_path = tmp_path / "cached_error.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "=B1/0"
    
    # We must patch writer to save cached error value
    import openpyxl.cell._writer as _cell_writer
    import openpyxl.worksheet._writer as _ws_writer
    from openpyxl.compat import safe_string
    from openpyxl.xml.functions import Element, SubElement
    
    original_write_cell = _cell_writer.etree_write_cell
    
    def mock_write_cell(xf, worksheet, cell, styled=None):
        if cell.coordinate == "A1":
            from openpyxl.cell._writer import _set_attributes
            _, attributes = _set_attributes(cell, styled)
            attributes['t'] = 'e'
            el = Element("c", attributes)
            formula = SubElement(el, 'f')
            formula.text = cell.value[1:]
            cell_content = SubElement(el, 'v')
            cell_content.text = safe_string("#DIV/0!")
            xf.write(el)
        else:
            original_write_cell(xf, worksheet, cell, styled)
            
    _cell_writer.etree_write_cell = mock_write_cell
    _ws_writer.write_cell = mock_write_cell
    
    try:
        wb.save(wb_path)
    finally:
        _cell_writer.etree_write_cell = original_write_cell
        _ws_writer.write_cell = original_write_cell
        
    scanner = WorkbookScanner(str(wb_path))
    res = scanner.scan()
    
    findings = [f for f in res["findings"] if f["rule_id"] == RULE_CACHED_ERROR]
    assert len(findings) == 1
    assert findings[0]["cell_address"] == "A1"
    assert findings[0]["severity"] == "critical"

def test_risk_score_calculation(tmp_path):
    wb_path = tmp_path / "risk_test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    # Let's add 1 critical (self ref) and 2 warnings (hardcoded number and external link)
    ws["A1"] = "=A1+1"  # critical: +30
    ws["B2"] = "=C2*0.08"  # warning: +10
    ws["D2"] = "=[Other.xlsx]Sheet1!A1"  # warning: +10
    # Total risk score should be 30 + 10 + 10 = 50
    wb.save(wb_path)
    
    scanner = WorkbookScanner(str(wb_path))
    res = scanner.scan()
    
    assert res["metadata"]["risk_score"] == 50
    assert res["metadata"]["status"] == "FAIL"  # fails because critical exists (even if score < 70)
