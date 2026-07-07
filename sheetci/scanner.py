import re
import datetime
from typing import Dict, List, Any, Optional
import openpyxl

# Severities
SEV_CRITICAL = "critical"
SEV_WARNING = "warning"
SEV_INFO = "info"

# Rule IDs
RULE_BROKEN_REF = "BROKEN_REF"
RULE_EXTERNAL_LINK = "EXTERNAL_LINK"
RULE_HIDDEN_SHEET = "HIDDEN_SHEET"
RULE_HARDCODED_NUMBER = "HARDCODED_NUMBER"
RULE_SELF_REFERENCE = "SELF_REFERENCE"
RULE_INCONSISTENT_FORMULA = "INCONSISTENT_FORMULA"
RULE_CACHED_ERROR = "CACHED_ERROR"

# Reference pattern for cell references (e.g., A1, $B$10, C$5)
# Ensures it is not followed by alphanumeric chars or a parenthesis (which indicates a function call like LOG10() or DEC2HEX())
REF_PATTERN = re.compile(r'(\$?)([A-Z]{1,3})(\$?)([0-9]+)(?![A-Z0-9_]|\s*\()', re.IGNORECASE)

# Standard Excel error values
EXCEL_ERRORS = {
    "#NULL!",
    "#DIV/0!",
    "#VALUE!",
    "#REF!",
    "#NAME?",
    "#NUM!",
    "#N/A",
}

def normalize_formula(formula: str, cell_row: int) -> str:
    """
    Normalizes a formula relative to the cell's row.
    Relative row references are converted to relative offsets (e.g. A2 -> A[0] if cell_row=2).
    """
    def replace_ref(match):
        col_abs = match.group(1)
        col = match.group(2)
        row_abs = match.group(3)
        row_str = match.group(4)
        
        if row_abs == '$':
            # Absolute row reference - keep as is
            return match.group(0)
        else:
            offset = int(row_str) - cell_row
            return f"{col_abs}{col}[{offset}]"
            
    return REF_PATTERN.sub(replace_ref, formula)

def extract_hardcoded_numbers(formula: str) -> List[float]:
    """
    Extracts hardcoded numeric constants from a formula string.
    Strips out string literals, sheet references, cell references/ranges, and function names first.
    """
    # 1. Strip string literals
    cleaned = re.sub(r'"[^"]*"', ' ', formula)
    cleaned = re.sub(r"'[^']*'", ' ', cleaned)
    
    # 2. Strip sheet references (e.g. 'Sheet 1'! or Sheet1!)
    cleaned = re.sub(r"'(?:[^']|'')+'!", ' ', cleaned)
    cleaned = re.sub(r"[a-zA-Z_0-9]+!", ' ', cleaned)
    
    # 3. Strip cell references & ranges
    cleaned = REF_PATTERN.sub(' ', cleaned)
    cleaned = re.sub(r'\b[A-Z]{1,3}:[A-Z]{1,3}\b', ' ', cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r'\b[0-9]+:[0-9]+\b', ' ', cleaned)
    
    # 4. Strip functions and identifiers (e.g. SUM, AVERAGE, DEC2HEX, LOG10)
    cleaned = re.sub(r'\b[A-Z_][A-Z0-9_\.]*\b', ' ', cleaned, flags=re.IGNORECASE)
    
    # 5. Find remaining numeric values
    numbers = re.findall(r'\b\d+(?:\.\d+)?\b', cleaned)
    
    results = []
    for num_str in numbers:
        try:
            val = float(num_str)
            # Ignore standard harmless constants: 0, 1, -1, 2
            # (since we look at unsigned values, 0, 1, 2 cover -0, -1, -2 as well)
            if val not in {0.0, 1.0, 2.0}:
                results.append(val)
        except ValueError:
            pass
    return results

class WorkbookScanner:
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.findings: List[Dict[str, Any]] = []
        self.metadata: Dict[str, Any] = {}
        
    def scan(self) -> Dict[str, Any]:
        self.findings = []
        
        # Load workbook twice:
        # 1. data_only=False to get formula definitions
        # 2. data_only=True to get cached evaluated values
        try:
            wb_formulas = openpyxl.load_workbook(self.filepath, data_only=False)
            wb_data = openpyxl.load_workbook(self.filepath, data_only=True)
        except Exception as e:
            raise ValueError(f"Failed to load Excel workbook {self.filepath}: {e}")
            
        total_sheets = len(wb_formulas.sheetnames)
        total_formulas = 0
        
        # Scan sheets
        for sheet_name in wb_formulas.sheetnames:
            ws_formulas = wb_formulas[sheet_name]
            ws_data = wb_data[sheet_name]
            
            # HIDDEN_SHEET check
            if ws_formulas.sheet_state in ("hidden", "veryHidden"):
                self.findings.append({
                    "sheet_name": sheet_name,
                    "cell_address": None,
                    "rule_id": RULE_HIDDEN_SHEET,
                    "severity": SEV_WARNING,
                    "explanation": f"Worksheet '{sheet_name}' is hidden or very hidden.",
                    "suggested_action": "Verify if the hidden sheet contains deprecated calculations or sensitive data."
                })
                
            # We will group formula cells by column for INCONSISTENT_FORMULA detection
            col_formulas: Dict[str, List[tuple]] = {}  # col_letter -> list of (cell, formula_str)
            
            for row in ws_formulas.iter_rows():
                for cell in row:
                    val = cell.value
                    if isinstance(val, str) and val.startswith("="):
                        total_formulas += 1
                        cell_address = cell.coordinate
                        formula_str = val
                        
                        # 1. BROKEN_REF check
                        if "#REF!" in formula_str.upper():
                            self.findings.append({
                                "sheet_name": sheet_name,
                                "cell_address": cell_address,
                                "rule_id": RULE_BROKEN_REF,
                                "severity": SEV_CRITICAL,
                                "explanation": f"Formula contains broken reference (#REF!): {formula_str}",
                                "suggested_action": "Fix the formula reference pointing to a deleted cell or range."
                            })
                            
                        # 2. EXTERNAL_LINK check
                        # Check for '[', '.xlsx', '.xls', 'http://', 'https://'
                        if any(x in formula_str.lower() for x in ["[", ".xlsx", ".xls", "http://", "https://"]):
                            self.findings.append({
                                "sheet_name": sheet_name,
                                "cell_address": cell_address,
                                "rule_id": RULE_EXTERNAL_LINK,
                                "severity": SEV_WARNING,
                                "explanation": f"Formula references an external workbook or URL: {formula_str}",
                                "suggested_action": "Ensure the external link is accessible, secure, and intended."
                            })
                            
                        # 4. HARDCODED_NUMBER check
                        hardcoded_nums = extract_hardcoded_numbers(formula_str)
                        if hardcoded_nums:
                            nums_str = ", ".join(str(n) for n in hardcoded_nums)
                            self.findings.append({
                                "sheet_name": sheet_name,
                                "cell_address": cell_address,
                                "rule_id": RULE_HARDCODED_NUMBER,
                                "severity": SEV_WARNING,
                                "explanation": f"Formula contains hardcoded numeric constants: {formula_str} (constants: {nums_str})",
                                "suggested_action": "Move hardcoded constants to input cells or parameters to make the model dynamic."
                            })
                            
                        # 5. SELF_REFERENCE check
                        col_letter = cell.column_letter
                        row_num = cell.row
                        self_ref_pattern = re.compile(rf"\b\$?{col_letter}\$?{row_num}\b", re.IGNORECASE)
                        if self_ref_pattern.search(formula_str):
                            self.findings.append({
                                "sheet_name": sheet_name,
                                "cell_address": cell_address,
                                "rule_id": RULE_SELF_REFERENCE,
                                "severity": SEV_CRITICAL,
                                "explanation": f"Formula contains a circular self-reference to its own cell: {formula_str}",
                                "suggested_action": "Refactor the formula to avoid referencing its own cell coordinate."
                            })
                            
                        # 7. CACHED_ERROR check
                        # Look up corresponding cell in ws_data
                        cached_val = ws_data[cell_address].value
                        if isinstance(cached_val, str) and cached_val in EXCEL_ERRORS:
                            self.findings.append({
                                "sheet_name": sheet_name,
                                "cell_address": cell_address,
                                "rule_id": RULE_CACHED_ERROR,
                                "severity": SEV_CRITICAL,
                                "explanation": f"Cell has a cached calculation error value: {cached_val}",
                                "suggested_action": "Review the cell's inputs or formula structure to resolve the evaluation error."
                            })
                            
                        # Collect for INCONSISTENT_FORMULA detection
                        col_letter = cell.column_letter
                        if col_letter not in col_formulas:
                            col_formulas[col_letter] = []
                        col_formulas[col_letter].append((cell, formula_str))
            
            # 6. INCONSISTENT_FORMULA check (column-by-column)
            for col_letter, cells_info in col_formulas.items():
                if len(cells_info) < 5:
                    continue  # Keep it conservative: only evaluate if >= 5 formulas in the column
                
                # Normalize all formulas
                normalized_list = []
                for cell, f_str in cells_info:
                    norm = normalize_formula(f_str, cell.row)
                    normalized_list.append(norm)
                    
                # Find majority normalized pattern
                pattern_counts: Dict[str, int] = {}
                for norm in normalized_list:
                    pattern_counts[norm] = pattern_counts.get(norm, 0) + 1
                    
                majority_pattern = max(pattern_counts, key=pattern_counts.get)
                majority_count = pattern_counts[majority_pattern]
                
                # Check if majority pattern represents >= 80%
                majority_percentage = majority_count / len(cells_info)
                if majority_percentage >= 0.8:
                    # Identify cells that deviate from majority
                    for (cell, f_str), norm in zip(cells_info, normalized_list):
                        if norm != majority_pattern:
                            self.findings.append({
                                "sheet_name": sheet_name,
                                "cell_address": cell.coordinate,
                                "rule_id": RULE_INCONSISTENT_FORMULA,
                                "severity": SEV_WARNING,
                                "explanation": f"Formula is inconsistent with neighboring cells in column {col_letter}: {f_str} (expected pattern similar to: {majority_pattern})",
                                "suggested_action": "Check if the formula was modified intentionally or copied down incorrectly."
                            })

        # Calculate risk score
        # critical: +30, warning: +10, info: +3, capped at 100
        risk_score = 0
        critical_count = 0
        warning_count = 0
        info_count = 0
        
        for finding in self.findings:
            sev = finding["severity"]
            if sev == SEV_CRITICAL:
                risk_score += 30
                critical_count += 1
            elif sev == SEV_WARNING:
                risk_score += 10
                warning_count += 1
            elif sev == SEV_INFO:
                risk_score += 3
                info_count += 1
                
        risk_score = min(risk_score, 100)
        
        # Pass/Fail conditions:
        # Fails if any critical finding exists OR risk score >= 70
        is_pass = (critical_count == 0) and (risk_score < 70)
        
        self.metadata = {
            "workbook_name": openpyxl.utils.escape.unescape(self.filepath.split("/")[-1]),
            "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "total_sheets": total_sheets,
            "total_formulas": total_formulas,
            "total_findings": len(self.findings),
            "risk_score": risk_score,
            "status": "PASS" if is_pass else "FAIL",
            "counts": {
                "critical": critical_count,
                "warning": warning_count,
                "info": info_count
            }
        }
        
        return {
            "metadata": self.metadata,
            "findings": self.findings
        }
