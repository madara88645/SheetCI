import os
import openpyxl
from openpyxl.compat import safe_string
from openpyxl.xml.functions import Element, SubElement, whitespace
import openpyxl.cell._writer as _cell_writer
import openpyxl.worksheet._writer as _ws_writer

# Dictionary to store custom cached values: (sheet_name, cell_coordinate) -> (cached_value_str, data_type_str)
CACHED_VALUES_MAPPING = {}

# Monkey-patch cell writing to support writing cached values
original_write_cell = _cell_writer.etree_write_cell

def patched_write_cell(xf, worksheet, cell, styled=None):
    key = (worksheet.title, cell.coordinate)
    if cell.data_type == 'f' and key in CACHED_VALUES_MAPPING:
        from openpyxl.cell._writer import _set_attributes
        _, attributes = _set_attributes(cell, styled)
        
        cached_val, data_type = CACHED_VALUES_MAPPING[key]
        if data_type:
            attributes['t'] = data_type
            
        el = Element("c", attributes)
        
        formula = SubElement(el, 'f')
        formula.text = cell.value[1:]  # strip leading '='
        
        cell_content = SubElement(el, 'v')
        cell_content.text = safe_string(cached_val)
        
        xf.write(el)
    else:
        original_write_cell(xf, worksheet, cell, styled)

_cell_writer.etree_write_cell = patched_write_cell
_ws_writer.write_cell = patched_write_cell

def create_broken_model(filepath: str):
    wb = openpyxl.Workbook()
    
    # 1. Main Sheet
    ws = wb.active
    ws.title = "Commissions"
    
    # Setup some data for inconsistent calculations
    ws["B2"] = 1000
    ws["C2"] = 0.05
    ws["D2"] = "=B2*C2"
    
    ws["B3"] = 2000
    ws["C3"] = 0.05
    ws["D3"] = "=B3*C3"
    
    ws["B4"] = 1500
    ws["C4"] = 0.05
    ws["D4"] = "=B4*C4"
    
    ws["B5"] = 3000
    ws["C5"] = 0.05
    ws["D5"] = "=B5*C5"
    
    ws["B6"] = 2500
    ws["C6"] = 0.05
    # Inconsistent formula in repeated column (D)
    ws["D6"] = "=B6*C6+100"
    
    ws["B7"] = 4000
    ws["C7"] = 0.05
    ws["D7"] = "=B7*C7"
    
    # Broken reference formula
    ws["E2"] = "=B2+#REF!"
    
    # Hardcoded number in formula
    ws["F2"] = "=B2*0.08"
    
    # External link-like formula
    ws["G2"] = "=[ExternalSource.xlsx]Sheet1!A1"
    
    # Direct self-reference
    ws["A1"] = "=A1+1"
    
    # Cached formula error (requires patch to write cached error string and set type 'e')
    ws["H2"] = "=B2/0"
    CACHED_VALUES_MAPPING[("Commissions", "H2")] = ("#DIV/0!", "e")
    
    # 2. Hidden Sheet
    hidden_ws = wb.create_sheet(title="HiddenArchive")
    hidden_ws.sheet_state = "hidden"
    hidden_ws["A1"] = "Secret Data"
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    print(f"Created broken commission model: {filepath}")

def create_clean_model(filepath: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commissions"
    
    ws["B2"] = 1000
    ws["C2"] = 0.05
    ws["D2"] = "=B2*C2"
    
    ws["B3"] = 2000
    ws["C3"] = 0.05
    ws["D3"] = "=B3*C3"
    
    ws["B4"] = 1500
    ws["C4"] = 0.05
    ws["D4"] = "=B4*C4"
    
    ws["B5"] = 3000
    ws["C5"] = 0.05
    ws["D5"] = "=B5*C5"
    
    ws["B6"] = 2500
    ws["C6"] = 0.05
    ws["D6"] = "=B6*C6"
    
    ws["B7"] = 4000
    ws["C7"] = 0.05
    ws["D7"] = "=B7*C7"
    
    # Harmless constant formulas (0, 1, 2)
    ws["E2"] = "=D2+0"
    ws["E3"] = "=D3*1"
    ws["E4"] = "=D4-1"
    ws["E5"] = "=D5/2"
    ws["E6"] = "=D6+0"
    ws["E7"] = "=D7+0"
    
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    print(f"Created clean model: {filepath}")

if __name__ == "__main__":
    create_broken_model("examples/broken-commission-model.xlsx")
    create_clean_model("examples/clean-model.xlsx")
